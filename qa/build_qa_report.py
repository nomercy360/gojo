from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD_FILL = PatternFill("solid", start_color="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
CELL_FONT = Font(name=FONT, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

PASS_FILL = PatternFill("solid", start_color="C6EFCE")
FAIL_FILL = PatternFill("solid", start_color="FFC7CE")
WARN_FILL = PatternFill("solid", start_color="FFEB9C")

wb = Workbook()

# ---------------- Sheet 1: Test Cases ----------------
ws = wb.active
ws.title = "Тест-кейсы"

headers = ["ID", "Область", "Что проверяем", "Шаги / запрос", "Ожидаемо",
           "Фактически", "Статус", "Приоритет", "Комментарий"]

# test data: id, area, what, steps, expected, actual, status, priority, comment
rows = [
    # AUTH & ROLES
    ["TC-01","Auth","Вход с верным паролем","POST /auth/sign-in/email, student.test / dev-password-123","200 + сессия","200","PASS","High","Базовый happy-path логина."],
    ["TC-02","Auth","Вход с неверным паролем","POST /auth/sign-in/email, wrongpass123","401","401","PASS","High","Негативный кейс, корректно отклоняется."],
    ["TC-03","Auth/Roles","Enum роли только student/admin","Проверка миграции 0013 в БД","enum = {student, admin}","student, admin","PASS","High","Роль teacher удалена, миграция применена без потери данных."],
    ["TC-04","Auth/Roles","teacher→admin при миграции","UPDATE role='admin' WHERE role='teacher'","0 записей teacher","0","PASS","Med","Ретро-конверсия отработала (проверено на сид-препод)."],
    ["TC-05","Roles","Гость не видит /admin/summary","GET /admin/summary без cookie","401/403","401","PASS","High","Требуется авторизация."],
    ["TC-06","Roles","Студент не видит /admin/summary","GET /admin/summary (student)","403","403","PASS","Crit","Разграничение доступа админ-панели."],
    ["TC-07","Roles","Админ видит /admin/summary","GET /admin/summary (admin)","200","200","PASS","High","Админ имеет доступ."],
    ["TC-08","Roles","Студент не видит /teacher/students","GET /teacher/students (student)","403","403","PASS","Crit","requireTeacher теперь = только admin."],
    ["TC-09","Roles","Админ видит /teacher/students","GET /teacher/students (admin)","200","200","PASS","High",""],
    ["TC-10","Roles","Студент не создаёт урок","POST /teacher/lessons (student)","403","403","PASS","Crit","teacher-роут защищён."],

    # PAYMENTS / PLAN
    ["TC-11","Payments","Провижн студента без planId","POST /admin/students без planId","400","400","PASS","High","Валидация обязательности тарифа."],
    ["TC-12","Payments","Провижн с несуществующим planId","POST /admin/students planId=nonexistent","400 'unknown plan'","400","PASS","High","Проверка whitelisted-планов."],
    ["TC-13","Payments","Назначение неизвестного плана","PATCH /teacher/students/:id/plan fake-plan","400","400","PASS","Med",""],
    ["TC-14","Payments","Студент назначает себе план (эскалация)","PATCH .../plan (student)","403","403","PASS","Crit","Защита от privilege escalation."],
    ["TC-15","Payments","payments/me со студентом с тарифом","GET /payments/me (student2)","assignedPlan=monthly-standard","OK","PASS","High","Тариф корректно отдаётся."],
    ["TC-16","Payments","payments/me студент без тарифа","GET /payments/me (student1)","assignedPlan=null","null","PASS","High","Фолбэк 'план не назначен'."],
    ["TC-17","Payments","Checkout чужого плана","POST /payments/checkout bundle-8 (assigned=monthly)","403 или 503","503*","PARTIAL","High","YooKassa не сконфигурирован (503) срабатывает РАНЬШЕ проверки плана. С кредами ранее давал 403 — см. Замечание F-05."],
    ["TC-18","Payments","Checkout админом","POST /payments/checkout (admin)","403 only students","403","PASS","Med",""],
    ["TC-19","Payments","payments/plans публичный","GET /payments/plans без auth","200","200","PASS","Low","Каталог тарифов публичен — ок."],

    # LIBRARY / RECORDINGS
    ["TC-20","Library","my-recordings без auth","GET /lessons/my-recordings","401","401","PASS","High","Список записей защищён."],
    ["TC-21","Library","my-recordings студент с записью","GET (student2)","1 запись N5","1 запись","PASS","High","Библиотека отдаёт реальную запись."],
    ["TC-22","Library","my-recordings студент без записей","GET (student1)","[] пусто","[]","PASS","High","Пустое состояние → плейсхолдер на дашборде."],
    ["TC-23","Library","Открытие записи → плеер","Дашборд→Библиотека→урок","video играет","играет","PASS","High","hero.mp4 (2.97MB) реально в Minio, плеер работает."],
    ["TC-24","Library","Файл записи в Minio","GET прямой URL .mp4","200, размер совпал","200 / 2974237 б","PASS","High","Байт-в-байт с исходником."],

    # PROFILE
    ["TC-25","Profile","Редактирование имя/фамилия","PATCH /users/me name","200 + сохранение","200","PASS","High","Имя разбивается/склеивается корректно."],
    ["TC-26","Profile","Имя >200 символов","PATCH name=201 симв.","400","400","PASS","Med","Валидация длины."],
    ["TC-27","Profile","Редирект на дашборд после сохранения","UI: Сохранить изменения","→ /dashboard","→ /dashboard","PASS","Med","redirect('/dashboard') отрабатывает."],
    ["TC-28","Profile","Тост 'Профиль сохранён'","UI после сохранения","показать тост","показан на /dashboard","PASS","Low","ИСПРАВЛЕНО (F-01): redirect '/dashboard?saved=1' + клиентский SavedToast. Проверено в браузере."],

    # FRONTEND ROUTES
    ["TC-29","Routes","Лендинг доступен","GET / ","200","200","PASS","High",""],
    ["TC-30","Routes","Кандзи-тренажёр удалён","GET /kanji","404","404","PASS","High","Роут удалён полностью."],
    ["TC-31","Routes","Protected → редирект гостя","GET /dashboard,/payments,/profile без auth","307→login","307","PASS","High","Все защищённые роуты редиректят."],
    ["TC-32","Routes","Дашборд авторизованному","GET /dashboard (student2)","200","200","PASS","High",""],
    ["TC-33","Routes","Навбар ЛК студента","UI /dashboard","только 'На сайт'+'Выйти'","OK","PASS","Med","Навигация вынесена в контент ЛК."],
    ["TC-34","UI/Landing","Hero CTA = 'Как это работает'","Лендинг hero","скролл на #how","OK","PASS","Low","Заменил 'ИИ тренажер'."],
    ["TC-35","UI/Landing","Instagram-кнопка","CTA-секция","'Instagram сообщество' → instagram.com/gojolearn","OK","PASS","Low",""],
    ["TC-36","UI/Landing","Telegram без дефиса","CTA + booking-modal","'Telegram сообщество'","OK","PASS","Low","Дефис убран в 2 местах."],
    ["TC-37","UI/Landing","Оранжевая линия перед футером убрана","section-cta::after","нет полоски","нет","PASS","Low",""],
    ["TC-38","UI/Landing","Фото Руслана = ~как Максим","section-mission photo","увеличено умеренно","112x143","PASS","Low","Итеративно подобран размер."],

    # SECURITY PROBES
    ["TC-39","Security","Материалы урока без auth","GET /lessons/:id/materials без cookie","401","401","PASS","Med","ИСПРАВЛЕНО (F-02): добавлен requireAuth + проверка брони/admin."],
    ["TC-40","Security","Прямой доступ к записи по Minio URL","GET Minio .mp4 без auth","(инфра)","200","WARN","High","App-слой закрыт (см. TC-45). Сам bucket public-read на уровне Minio — инфра-задача F-03, не код."],
    ["TC-41","Security","dev-login может выдать admin","POST /dev-auth/dev-login role=admin","gated + warning","200 (dev)","PASS","High","Митигировано (F-04): в prod гейтится + лог-warning при старте."],
    ["TC-42","Security","404 несуществующий урок","GET /lessons/<random uuid>","404","404","PASS","Low","Нет утечки через ошибку."],

    # RE-TEST после фиксов
    ["TC-43","Security","F-02 материалы: не забронирован → 403","GET materials (student1 не booked)","403","403","PASS","Med","Ре-тест фикса F-02."],
    ["TC-44","Security","F-02 материалы: забронирован → 200","GET materials (student2 booked)","200","200","PASS","Med","Легитимный доступ не сломан."],
    ["TC-45","Security","F-03 recordingUrl скрыт неавторизованным","GET /lessons/:id (no auth / not booked)","recordingUrl=null","null","PASS","High","Ре-тест F-03: гость и не-забронированный не получают URL записи."],
    ["TC-46","Security","F-03 recordingUrl виден booked/admin","GET /lessons/:id (student2 / admin)","URL present","present","PASS","High","Видео у забронированного играет (проверено в браузере)."],
]

# title
ws.merge_cells("A1:I1")
ws["A1"] = "Gojo Learn — QA тест-кейсы и результаты"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 24

ws.merge_cells("A2:I2")
ws["A2"] = "Дата: 2026-07-10 · Окружение: локальное (Docker infra + dev api/web) · Тестировщик: QA Manager (Claude) · Раунд 2: после фиксов F-01…F-04"
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

hr = 3
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = CENTER; cell.border = BORDER

for i, r in enumerate(rows, start=hr+1):
    for c, val in enumerate(r, 1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.font = CELL_FONT; cell.border = BORDER; cell.alignment = WRAP
    st = ws.cell(row=i, column=7)
    st.alignment = CENTER
    status = r[6]
    if status == "PASS": st.fill = PASS_FILL
    elif status == "FAIL": st.fill = FAIL_FILL
    else: st.fill = WARN_FILL  # PARTIAL / WARN

widths = [8, 13, 26, 34, 24, 22, 9, 9, 42]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A{hr}:I{hr+len(rows)}"

last = hr + len(rows)

# ---------------- Sheet 2: Summary ----------------
s2 = wb.create_sheet("Сводка")
s2.merge_cells("A1:C1")
s2["A1"] = "Сводка по результатам"
s2["A1"].font = TITLE_FONT
s2.row_dimensions[1].height = 24

data_range = f"'Тест-кейсы'!$G${hr+1}:$G${last}"
metrics = [
    ("Всего тест-кейсов", f'=COUNTA({data_range})'),
    ("PASS", f'=COUNTIF({data_range},"PASS")'),
    ("FAIL", f'=COUNTIF({data_range},"FAIL")'),
    ("PARTIAL", f'=COUNTIF({data_range},"PARTIAL")'),
    ("WARN", f'=COUNTIF({data_range},"WARN")'),
    ("% PASS", f'=COUNTIF({data_range},"PASS")/COUNTA({data_range})'),
]
r0 = 3
for i, (label, formula) in enumerate(metrics):
    lc = s2.cell(row=r0+i, column=1, value=label)
    lc.font = Font(name=FONT, bold=True, size=10); lc.border = BORDER
    vc = s2.cell(row=r0+i, column=2, value=formula)
    vc.font = CELL_FONT; vc.border = BORDER; vc.alignment = CENTER
    if label == "% PASS": vc.number_format = "0.0%"
    if label == "PASS": vc.fill = PASS_FILL
    if label == "FAIL": vc.fill = FAIL_FILL
    if label in ("PARTIAL","WARN"): vc.fill = WARN_FILL

# by-area breakdown
s2.cell(row=r0, column=4, value="Область").font = HEAD_FONT
s2.cell(row=r0, column=4).fill = HEAD_FILL
s2.cell(row=r0, column=5, value="Кейсов").font = HEAD_FONT
s2.cell(row=r0, column=5).fill = HEAD_FILL
areas = ["Auth","Roles","Payments","Library","Profile","Routes","UI/Landing","Security"]
area_range = f"'Тест-кейсы'!$B${hr+1}:$B${last}"
for i, a in enumerate(areas):
    ac = s2.cell(row=r0+1+i, column=4, value=a); ac.font=CELL_FONT; ac.border=BORDER
    cc = s2.cell(row=r0+1+i, column=5, value=f'=COUNTIF({area_range},"{a}*")')
    cc.font=CELL_FONT; cc.border=BORDER; cc.alignment=CENTER

for col, w in zip("ABCDE", [20, 12, 4, 14, 10]):
    s2.column_dimensions[col].width = w

s2.cell(row=r0+len(metrics)+2, column=1,
        value="Вердикт: релиз готов. F-01/F-02/F-04 исправлены и перепроверены; F-03 закрыт на уровне приложения (recordingUrl гейтится), приватность Minio-bucket — отдельная инфра-задача. Блокеров нет.").font = Font(name=FONT, italic=True, size=9, color="595959")
s2.merge_cells(start_row=r0+len(metrics)+2, start_column=1, end_row=r0+len(metrics)+2, end_column=5)

# ---------------- Sheet 3: Findings ----------------
s3 = wb.create_sheet("Замечания")
s3.merge_cells("A1:E1")
s3["A1"] = "Замечания и рекомендации"
s3["A1"].font = TITLE_FONT
s3.row_dimensions[1].height = 24

fh = ["ID","Severity","Заголовок","Описание","Рекомендация","Статус"]
findings = [
    ["F-01","Low (регресс)","Тост 'Профиль сохранён' не показывается",
     "После добавления redirect('/dashboard') в updateProfileAction экшен больше не возвращал {ok:true}, поэтому useEffect с toast.success был мёртв.",
     "redirect('/dashboard?saved=1') + клиентский компонент SavedToast на дашборде, который показывает тост и чистит URL.","ИСПРАВЛЕНО ✓ (проверено в браузере)"],
    ["F-02","Medium","GET /lessons/:id/materials открыт без авторизации",
     "Эндпоинт материалов урока отдавал список с прямыми fileUrl любому без сессии. Утечка ссылок на приватные материалы. Пред-существующее.",
     "Добавлен requireAuth + hasLessonContentAccess (admin / владелец урока / забронированный студент).","ИСПРАВЛЕНО ✓ (401/403/200 по ролям)"],
    ["F-03","High","recordingUrl раздавался неавторизованным + bucket public-read",
     "App: GET /lessons/:id возвращал recordingUrl любому. Инфра: сам файл в Minio (gojo-dev) world-readable по прямому URL.",
     "APP-СЛОЙ ИСПРАВЛЕН: recordingUrl теперь отдаётся только booked/owner/admin (иначе null). ИНФРА-ЗАДАЧА (остаётся): сделать bucket/prefix materials приватным + presigned URL; avatars/ оставить public. Не сделано в этой сессии сознательно — ломает рендер аватаров и требует настройки подписи за Caddy-прокси в prod.","ЧАСТИЧНО: app ✓, инфра — TODO"],
    ["F-04","Medium","dev-login может выдать admin-сессию",
     "POST /dev-auth/dev-login role=admin создаёт админа. Гейтится NODE_ENV===production && !ALLOW_DEV_LOGIN.",
     "Добавлен лог-warning при старте API, если NODE_ENV=production и ALLOW_DEV_LOGIN=true.","ИСПРАВЛЕНО ✓ (warning добавлен)"],
    ["F-05","Info","Порядок проверок в checkout",
     "В /payments/checkout проверка yookassaConfigured() (503) идёт РАНЬШЕ проверки соответствия плана (403). С кредами 403 подтверждён ранее.",
     "Не баг. Опционально: поставить проверку плана перед проверкой провайдера.","НЕ ТРЕБУЕТ ДЕЙСТВИЙ"],
    ["F-06","Info","Тестовые данные в БД",
     "Создан урок 'Тестовая запись урока — N5' + бронь student2 + hero.mp4 как запись. Аккаунт provisioned.test@ имеет случайный пароль.",
     "Перед прод-деплоем прогнать чистую миграцию на пустой БД; тестовые сиды не коммитить.","К СВЕДЕНИЮ"],
]
for c, h in enumerate(fh, 1):
    cell = s3.cell(row=3, column=c, value=h)
    cell.fill=HEAD_FILL; cell.font=HEAD_FONT; cell.alignment=CENTER; cell.border=BORDER
for i, f in enumerate(findings, start=4):
    for c, val in enumerate(f, 1):
        cell = s3.cell(row=i, column=c, value=val)
        cell.font=CELL_FONT; cell.border=BORDER; cell.alignment=WRAP
    sev = s3.cell(row=i, column=2)
    if sev.value.startswith("High"): sev.fill = FAIL_FILL
    elif sev.value.startswith("Medium") or sev.value.startswith("Low"): sev.fill = WARN_FILL
    else: sev.fill = PatternFill("solid", start_color="DDEBF7")
    stc = s3.cell(row=i, column=6)
    if stc.value.startswith("ИСПРАВЛЕНО"): stc.fill = PASS_FILL
    elif stc.value.startswith("ЧАСТИЧНО"): stc.fill = WARN_FILL

for col, w in zip("ABCDEF", [8, 15, 28, 52, 46, 24]):
    s3.column_dimensions[col].width = w
s3.freeze_panes = "A4"

wb.save("C:/Users/tammu/gojo/qa/QA_Report_Gojo.xlsx")
print("saved")
